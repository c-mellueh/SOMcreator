import os

import openpyxl
from lxml import etree, builder
from openpyxl import load_workbook,Workbook
from openpyxl.worksheet.worksheet import  Worksheet

from .. import classes


def _create_sheet(obj:classes.Object, workbook:Workbook,name):
    sheet = workbook.create_sheet(name)
    attributes = set()
    for property_set in obj.property_sets:
        for attribute in property_set.attributes:
            attributes.add(attribute.name)
    for i,attrib_name in enumerate(sorted(attributes),start=1):
        sheet.cell(1,i).value = attrib_name


def create_mapping(src_path: str, dest_path: str, project:classes.Project) -> None:


    export_wb = Workbook()
    export_wb.active.title="Hilfe"
    wb = load_workbook(src_path)
    sheet:Workbook = wb.active
    important_rows = [row for i, row in enumerate(sheet.rows) if row[2].value is not None and i != 0]
    object_dict = {obj.ident_attrib.value[0]: obj for obj in project.objects if not obj.is_concept}

    for row in important_rows:
        bauteil_bez_card, bauteil_bez_2, bauteilklass = map(lambda x: x.value, row)
        obj = object_dict.get(bauteilklass)
        if obj is None:
            print(f"BauteilKlassifikation '{bauteilklass}' not found")
            continue
        _create_sheet(obj,export_wb,bauteil_bez_card)
    export_wb.save(dest_path)
