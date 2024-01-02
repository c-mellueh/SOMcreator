from __future__ import annotations

import json
import logging
import os.path
import re
import shutil
import tempfile
from typing import Iterator

import openpyxl
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from .. import classes
from ..Template import IFC_4_1
from ..constants import value_constants

IDENT_PSET_NAME = "Allgemeine Eigenschaften"
IDENT_ATTRIB_NAME = "bauteilKlassifikation"


def split_string(text: str) -> list[str] | None:
    if text is None:
        return []
    text = re.findall(r'[^,;]+', text)
    for i, item in enumerate(text):
        if "(" in item:
            item = item.split("(")
            text[i] = item[0]
        text[i] = text[i].strip()

    return text


def find_by_abbreviation(abbreviation: str) -> ExcelBlock | None:
    excel_block_dict: dict[str, ExcelBlock] = {block.abbreviation.upper(): block for block in ExcelBlock}
    return excel_block_dict.get(abbreviation.upper())


def autoadjust_column_widths(sheet: Worksheet) -> None:
    for i in range(len(list(sheet.columns))):
        column_letter = get_column_letter(i + 1)
        column = sheet[column_letter]
        width = max([len(cell.value) for cell in column if cell.value is not None], default=2)
        sheet.column_dimensions[column_letter].width = width


class ExcelIterator(type):
    def __iter__(cls: ExcelBlock) -> Iterator[ExcelBlock]:
        return iter(cls._registry)


class ExcelBlock(metaclass=ExcelIterator):
    _registry: list[ExcelBlock] = list()

    def __init__(self, base_cell: Cell, sheet: Worksheet):
        self._registry.append(self)
        self.base_cell = base_cell
        self.sheet = sheet
        self.pset: classes.PropertySet | None = None
        self.object: classes.Object | None = None

    @property
    def ident_value(self) -> str | None:
        cell = self.sheet.cell(row=self.base_cell.row,
                               column=self.base_cell.column + 2)
        return cell.value

    @property
    def name(self) -> str:
        cell = self.sheet.cell(row=self.base_cell.row,
                               column=self.base_cell.column + 1)
        return cell.value

    @property
    def abbreviation(self) -> str:
        cell = self.sheet.cell(row=self.base_cell.row + 1,
                               column=self.base_cell.column + 1)
        return cell.value.upper()

    @property
    def is_predefined_pset(self) -> bool:
        if self.ident_value is None:
            return True
        else:
            return False

    @property
    def entry(self) -> Cell:
        if self.is_predefined_pset:
            return self.sheet.cell(row=self.base_cell.row + 4, column=self.base_cell.column)
        else:
            return self.sheet.cell(row=self.base_cell.row + 5, column=self.base_cell.column)

    @property
    def aggregates(self) -> set[ExcelBlock] | None:
        if self.is_predefined_pset:
            return None

        child_cell = self.sheet.cell(row=self.base_cell.row + 3, column=self.base_cell.column + 1)
        child_string_list = split_string(child_cell.value)
        children = set()
        for abbrev in child_string_list:
            if abbrev != "-":
                block = find_by_abbreviation(abbrev)
                if block is not None:
                    children.add(block)
                else:
                    logging.error(f"[{self.name}] abbreviation '{abbrev}' doesn't exist!")
        return children

    @property
    def parent_classes(self) -> set[ExcelBlock] | None:
        parents = set()

        parent_text = self.sheet.cell(self.base_cell.row + 2, self.base_cell.column + 1).value
        for abbrev in split_string(parent_text):
            if abbrev != "-":
                parent_block = find_by_abbreviation(abbrev)
                if parent_block is not None:
                    if parent_block.name != IDENT_PSET_NAME:
                        parents.add(parent_block)
                        parents = set.union(parent_block.parent_classes, parents)
                else:

                    logging.warning(
                        f"[{self.name}] Elternklasse: Kürzel {abbrev.upper()} existiert nicht!")
        return parents

    @property
    def inherits(self) -> set[ExcelBlock]:

        def is_child(value) -> bool:
            last_num = value.split(".")[-1]

            if not last_num.isdigit():
                return False
            last_num = int(last_num)
            if last_num >= 100 and last_num % 100 == 0:
                return True
            else:
                return False

        inherit_blocks = set()

        if self.is_predefined_pset:
            return inherit_blocks

        for block in ExcelBlock:
            if not block.is_predefined_pset:
                parent_txt = block.ident_value.split(".")[:-1]
                parent_txt = ".".join(parent_txt)
                if parent_txt == self.ident_value and is_child(block.ident_value):
                    inherit_blocks.add(block)

        return inherit_blocks

    def create_attributes(self) -> set[classes.Attribute]:
        """
           Iterate over Attributes
           Create Them and find special Datatypes
           """

        def transform_value_types(attribute_name: str, value: str) -> (str, bool):
            if value is not None:
                if value.lower() in ["string", "str"]:
                    data_type = value_constants.LABEL
                elif value.lower() in ["double"]:
                    data_type = value_constants.REAL
                elif value.lower() in ["boolean", "bool"]:
                    data_type = value_constants.BOOLEAN
                elif value.lower() in ["int", "integer"]:
                    data_type = value_constants.INTEGER
                else:
                    text = f"{self.name}: Datatype '{value}' of Attribute {attribute_name} can't be interpreted. Use {value_constants.LABEL} instead "
                    logging.info(text)
                    data_type = value_constants.LABEL
            else:
                data_type = value_constants.LABEL

            return data_type

        entry = self.entry

        cell_list = set(block.base_cell for block in ExcelBlock)
        attributes: set[classes.Attribute] = set()

        while entry.value is not None and entry not in cell_list and entry.value != "-":
            data_type_text = self.sheet.cell(row=entry.row, column=entry.column + 2).value
            data_type = transform_value_types(entry.value, data_type_text)
            attribute_name: str = entry.value
            description = self.sheet.cell(row=entry.row, column=entry.column + 1).value
            optional = False
            if attribute_name.startswith("*"):
                optional = True
                attribute_name = attribute_name[1:]

            attribute = classes.Attribute(self.pset, attribute_name, [],
                                          value_constants.VALUE_TYPE_LOOKUP[value_constants.LIST],
                                          data_type=data_type, optional=optional, description=description)
            attribute.revit_name = attribute_name
            attributes.add(attribute)

            entry = self.sheet.cell(row=entry.row + 1, column=entry.column)

        return attributes

    def create_predefined_pset(self) -> None:
        optional = False
        pset_name = self.name
        if self.name.startswith("*"):
            optional = True
            pset_name = pset_name[1:]
        self.pset = classes.PropertySet(pset_name, optional=optional)
        self.pset.attributes = self.create_attributes()

    def create_object(self) -> classes.Object:
        pset_name = self.name
        if self.name.startswith("*"):
            pset_name = pset_name[1:]

        self.pset = classes.PropertySet(pset_name)
        attributes = self.create_attributes()
        if not attributes:
            self.pset.delete()
            self.pset = None

        predef_psets: dict[str, ExcelBlock] = {block.name: block for block in ExcelBlock if block.is_predefined_pset}
        parent_pset = predef_psets.get(IDENT_PSET_NAME)

        ident_pset = parent_pset.pset.create_child(IDENT_PSET_NAME)
        ident_attrib = ident_pset.get_attribute_by_name(IDENT_ATTRIB_NAME)
        ident_attrib.value = [self.ident_value]
        optional = False
        obj_name = self.name
        if self.name.startswith("*"):
            optional = True
            obj_name = obj_name[1:]
        obj = classes.Object(obj_name, ident_attrib, ifc_mapping=self.ifc_mapping(), optional=optional, )
        if attributes:
            obj.add_property_set(self.pset)
        obj.add_property_set(ident_pset)
        obj.abbreviation = self.abbreviation

        return obj

    def ifc_mapping(self) -> set[str]:
        if self.is_predefined_pset:
            return set()

        cell = self.sheet.cell(row=self.base_cell.row + 4, column=self.base_cell.column + 2)
        value: str = cell.value
        if value is None:
            logging.error(f"[{self.name}]: no IFC Mapping")
            return {"IfcBuildingElementProxy"}

        string_list = value.split("/")
        string_list = [string.strip() for string in string_list]

        for string in string_list:
            if string not in IFC_4_1:
                logging.info(f"[{self.name}]: '{string}' not in IFC 4.1 Specification")
        return set(string_list)


def _create_blocks(sheet) -> None:
    """create Excel Blocks"""
    excel_blocks = set()

    row: tuple[Cell]
    for row in sheet:
        for cell in row:
            if cell.value is not None:
                text = cell.value.strip()
                if text in ["name", "name:"]:
                    if sheet.cell(cell.row + 1, cell.column).value == "Kürzel":
                        excel_blocks.add(ExcelBlock(cell, sheet))
                    else:
                        logging.error(f"{sheet.cell(cell.row + 1, cell.column)} hat den Wert 'name'")


def _create_items() -> None:
    """create Objects and PropertySets"""
    predef_psets = [block for block in ExcelBlock if block.is_predefined_pset]
    objects = [block for block in ExcelBlock if not block.is_predefined_pset]

    for predef_pset_block in predef_psets:
        predef_pset_block.create_predefined_pset()

    for object_block in objects:
        obj = object_block.create_object()
        object_block.object = obj

    for block in objects:
        for pset in [block.pset for block in block.parent_classes if block.pset is not None]:
            new_pset = pset.create_child(pset.name)
            block.object.add_property_set(new_pset)

    # for SOM of Deutsche Bahn
    for obj in classes.Object:
        prop = obj.get_property_set_by_name("Allgemeine Eigenschaften")
        if prop is None:
            continue
        bn = prop.get_attribute_by_name("bauteilName")
        if bn is not None:
            bn.value = [obj.name]


def _build_object_tree() -> None:
    tree_dict: dict[str, classes.Object] = {obj.ident_attrib.value[0]: obj for obj in classes.Object}

    for ident, item in tree_dict.items():
        ident_list = ident.split(".")[:-1]
        parent_obj = tree_dict.get(".".join(ident_list))

        if parent_obj is not None:
            parent_obj.add_child(item)


def _build_aggregations() -> None:
    def get_root_blocks() -> set[ExcelBlock]:
        children = set()
        for e_block in ExcelBlock:  # find all child blocks
            if not e_block.is_predefined_pset:
                children = set.union(e_block.aggregates, children)
                children = set.union(e_block.inherits, children)

        r_blocks = set()  # root blocks
        for e_block in ExcelBlock:
            if e_block not in children and not e_block.is_predefined_pset:
                r_blocks.add(e_block)

        return r_blocks

    def link_child_nodes(aggregation: classes.Aggregation, block: ExcelBlock) -> None:
        aggregate_list = block.aggregates
        inherit_list = block.inherits

        for aggregate_block in aggregate_list:
            if aggregate_block.name != block.name:
                if aggregate_block.is_predefined_pset:
                    logging.error(f"[{block.name}] can't aggregate to {aggregate_block.name}")
                else:
                    child_aggreg = classes.Aggregation(aggregate_block.object)
                    relationship = value_constants.AGGREGATION
                    if aggregate_block in inherit_list:
                        relationship += value_constants.INHERITANCE

                    aggregation.add_child(child_aggreg, relationship)
                    link_child_nodes(child_aggreg, aggregate_block)
            else:
                logging.warning(f"[{aggregation.name}] recursive aggregation")

        for inherit_block in inherit_list:
            if inherit_block not in aggregate_list:
                child_aggreg = classes.Aggregation(inherit_block.object)
                aggregation.add_child(child_aggreg, value_constants.INHERITANCE)
                link_child_nodes(child_aggreg, inherit_block)

    root_blocks = get_root_blocks()

    for block in root_blocks:
        aggregation = classes.Aggregation(block.object)
        link_child_nodes(aggregation, block)


def open_file(path: str, ws_name: str) -> None:
    # TODO: add request for Identification Attribute
    logging.warning(f"Opening Excel files is deprecated and wont have newer features")
    with tempfile.TemporaryDirectory() as tmpdirname:
        new_path = os.path.join(tmpdirname, "excel.xlsx")
        shutil.copy2(path, new_path)
        book = openpyxl.load_workbook(new_path)
        if ws_name not in book.sheetnames:
            logging.error("Worksheet Name not in Workbook")
            return
        sheet = book[ws_name]

        _create_blocks(sheet)
        _create_items()
        _build_object_tree()
        _build_aggregations()

    ExcelBlock._registry = list()


def create_abbreviation_json(excel_path: str, ws_name: str, export_path: str = None) -> dict[str:[str, str]]:
    ExcelBlock._registry = list()
    with tempfile.TemporaryDirectory() as tmpdirname:
        new_path = os.path.join(tmpdirname, "excel.xlsx")
        shutil.copy2(excel_path, new_path)
        book = openpyxl.load_workbook(new_path)
        if ws_name not in book.sheetnames:
            logging.error("Worksheet Name not in Workbook")
            return
        sheet = book[ws_name]

        _create_blocks(sheet)
        _create_items()
        _build_object_tree()
        _build_aggregations()

        d = {block.abbreviation: [block.ident_value, block.name] for block in ExcelBlock if
             block.ident_value is not None}
        if export_path is not None:
            with open(export_path, "w") as file:
                json.dump(d, file, indent=2)
        return d


def export(project: classes.Project, path: str, mapping_dict: dict[str, str] = {}) -> None:
    if not os.path.exists(os.path.dirname(path)):
        raise FileNotFoundError(f"path {os.path.dirname(path)} DNE")

    NAME = "name"
    OBJECTS = "objects"
    TABLE_STYLE = "TableStyleLight1"
    OPTIONAL_FONT = styles.Font(color="4e6ec0")

    def fill_main_sheet(sheet: Worksheet) -> None:
        sheet.title = "Uebersicht"
        sheet.cell(1, 1).value = "bauteilName"
        sheet.cell(1, 2).value = "bauteilKlassifikation"
        row = 1
        for row, obj in enumerate(sorted(project.objects), start=2):
            sheet.cell(row, 1).value = obj.name
            sheet.cell(row, 2).value = str(obj.ident_value)
            if obj.optional:
                sheet.cell(row, 1).font = OPTIONAL_FONT
                sheet.cell(row, 2).font = OPTIONAL_FONT
        table_start = sheet.cell(1, 1).coordinate
        table_end = sheet.cell(row, 2).coordinate
        table_range = f"{table_start}:{table_end}"
        table = Table(displayName="Uebersicht", ref=table_range)
        style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)
        autoadjust_column_widths(sheet)

    def filter_to_sheets() -> dict:
        d = {obj.ident_value: {NAME: obj.name, OBJECTS: []} for obj in project.objects if
             len(obj.ident_value.split(".")) == 1}
        for ident, name in mapping_dict.items():
            d[ident] = {NAME: name, OBJECTS: []}

        for obj in project.objects:
            group = obj.ident_value.split(".")[0]
            d[group][OBJECTS].append(obj)

        d["son"] = {NAME: "Sonstige", OBJECTS: []}
        for group_name, group in list(d.items()):
            objects = group[OBJECTS]
            if len(objects) < 3:
                d["son"][OBJECTS] += objects
                del d[group_name]
        return d

    def create_object_entry(obj: classes.Object, sheet, start_row, start_column, table_index):
        if obj.optional:
            font_style = OPTIONAL_FONT
        else:
            font_style = styles.Font()

        sheet.cell(start_row, start_column).value = "bauteilName"
        sheet.cell(start_row, start_column + 1).value = obj.name

        sheet.cell(start_row + 1, start_column).value = "bauteilKlassifikation"
        sheet.cell(start_row + 1, start_column + 1).value = obj.ident_value

        sheet.cell(start_row + 2, start_column).value = "Kürzel"
        sheet.cell(start_row + 2, start_column + 1).value = str(obj.abbreviation)

        sheet.cell(start_row + 3, start_column).value = "Property"
        sheet.cell(start_row + 3, start_column + 1).value = "Propertyset"
        sheet.cell(start_row + 3, start_column + 2).value = "Beispiele / Beschreibung"
        sheet.cell(start_row + 3, start_column + 3).value = "Datentyp"

        for i in range(0, 4):
            for k in range(0, 4):
                sheet.cell(start_row + i, start_column + k).font = font_style
        draw_border(sheet, [start_row, start_row + 2], [start_column, start_column + 4])
        fill_grey(sheet, [start_row, start_row + 2], [start_column, start_column + 3])

        pset_start_row = start_row + 4
        index = 0
        for property_set in sorted(obj.property_sets):
            for attribute in sorted(property_set.attributes):
                sheet.cell(pset_start_row + index, start_column).value = attribute.name
                sheet.cell(pset_start_row + index, start_column + 1).value = property_set.name
                sheet.cell(pset_start_row + index, start_column + 2).value = attribute.description
                sheet.cell(pset_start_row + index, start_column + 3).value = attribute.data_type
                if attribute.optional:
                    sheet.cell(pset_start_row + index, start_column).font = OPTIONAL_FONT
                    sheet.cell(pset_start_row + index, start_column + 1).font = OPTIONAL_FONT
                    sheet.cell(pset_start_row + index, start_column + 2).font = OPTIONAL_FONT
                index += 1

        table_start = sheet.cell(pset_start_row - 1, start_column).coordinate
        table_end = sheet.cell(pset_start_row + index - 1, start_column + 3).coordinate
        table_range = f"{table_start}:{table_end}"
        table = Table(displayName=f"Tabelle_{str(table_index).zfill(5)}", ref=table_range)
        style = TableStyleInfo(name=TABLE_STYLE, showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        sheet.add_table(table)

    def draw_border(sheet: Worksheet, row_range: [int, int], column_range: [int, int]):
        for row in range(row_range[0], row_range[1] + 1):
            for column in range(column_range[0], column_range[1] + 1):
                left_side = styles.Side(border_style="none", color="FF000000")
                right_side = styles.Side(border_style="none", color="FF000000")
                top_side = styles.Side(border_style="none", color="FF000000")
                bottom_side = styles.Side(border_style="none", color="FF000000")
                if column == column_range[0]:
                    left_side = styles.Side(border_style="thick", color="FF000000")

                if column == column_range[1]:
                    right_side = styles.Side(border_style="thick", color="FF000000")

                if row == row_range[0]:
                    top_side = styles.Side(border_style="thick", color="FF000000")
                if row == row_range[1]:
                    bottom_side = styles.Side(border_style="thick", color="FF000000")
                sheet.cell(row, column).border = styles.Border(left=left_side, right=right_side, top=top_side,
                                                               bottom=bottom_side)

    def fill_grey(sheet: Worksheet, row_range: [int, int], column_range: [int, int]):
        fill = styles.PatternFill(fill_type="solid", start_color="d9d9d9")
        for row in range(row_range[0], row_range[1] + 1):
            for column in range(column_range[0], column_range[1] + 1):
                sheet.cell(row, column).fill = fill

    workbook = Workbook()

    sheet_main = workbook.active
    fill_main_sheet(sheet_main)

    sheet_dict = filter_to_sheets()
    table_counter = 1
    for ident, d in sheet_dict.items():
        obj_name = d[NAME]
        objects = d[OBJECTS]
        work_sheet = workbook.create_sheet(f"{obj_name} ({ident})")
        for counter, obj in enumerate(sorted(objects)):
            column = 1 + counter * 5
            create_object_entry(obj, work_sheet, 1, column, table_counter)
            table_counter += 1
        autoadjust_column_widths(work_sheet)
    workbook.save(path)
